# =============================================================================
#  parse_cupc.py  -  Parse CUPC headcount workbooks into UPP% + LCFF+ flag.
# =============================================================================
#  PART OF: the UC Admissions x School-Merit panel build.  See README.md
#
#  WHAT IT DOES
#    Reads the CUPC (Unduplicated Pupil Count) Excel workbooks and extracts,
#    per school per year, the 9-12 Unduplicated Pupil Percentage (UPP%) and the
#    "75%+" LCFF+ flag. UPP/LCFF+ is the documented UC equity lever.
#
#  *** RUN FROM:  the project root  ***
#  HOW TO RUN     python build_scripts_documented/parse_cupc.py
#
#  INPUTS   Headcounts/cupc*-912.xls*   (one workbook per year, 2016-17..2025-26)
#           (.xls read via xlrd, .xlsx via openpyxl)
#  OUTPUTS  build_outputs/upp_lcff.csv
#           columns: cds14, cupc_year, enroll_9_12, upp_pct, lcff_plus_flag
#
#  GOTCHAS / NOTES
#    - Column POSITIONS shift between years, so columns are found by HEADER NAME,
#      not by index.
#    - cupc_year comes from the filename: cupcYYZZ -> 20YY-20ZZ.
#    - Downstream (merge / then_vs_now / yearly_trend) read this from
#      Panel Build .../components/ -- copy upp_lcff.csv into components/ after
#      running (see README "copy steps").
#    - Requires the Python packages openpyxl AND xlrd.
# =============================================================================

import glob,os,csv,re,sys
def find(hdr,*subs):
    for j,c in enumerate(hdr):
        cl=str(c).lower().replace("\n"," ")
        if all(s in cl for s in subs): return j
    return None
def sheet_name(names):
    for s in names:
        if "upc" in s.lower() and ("9-12" in s or "912" in s.replace(" ","")): return s
    for s in names:
        if "upc" in s.lower(): return s
    return names[1] if len(names)>1 else names[0]
def rows_xlsx(fp):
    import openpyxl
    wb=openpyxl.load_workbook(fp,read_only=True,data_only=True); ws=wb[sheet_name(wb.sheetnames)]
    for row in ws.iter_rows(values_only=True): yield list(row)
def rows_xls(fp):
    import xlrd
    b=xlrd.open_workbook(fp); s=b.sheet_by_name(sheet_name(b.sheet_names()))
    for i in range(s.nrows): yield [s.cell_value(i,c) for c in range(s.ncols)]
out=[]
for fp in sorted(glob.glob("Headcounts/cupc*-912.xls*")):
    m=re.search(r'cupc(\d{2})(\d{2})-912',os.path.basename(fp)); yr=f"20{m.group(1)}-20{m.group(2)}"
    gen=rows_xls(fp) if fp.lower().endswith(".xls") else rows_xlsx(fp)
    cols=None; n=0
    for row in gen:
        if cols is None:
            if find(row,"county","code") is not None and find(row,"school","code") is not None:
                cols=dict(C=find(row,"county","code"),D=find(row,"district","code"),S=find(row,"school","code"),
                          E=find(row,"total","enrollment"),U=find(row,"unduplicated","percentage"),F=find(row,"75%"))
            continue
        sc=row[cols["S"]]
        if sc in (None,""): continue
        school=str(sc).split(".")[0].strip().zfill(7)
        if school=="0000000": continue
        cds=str(row[cols["C"]]).split(".")[0].strip().zfill(2)+str(row[cols["D"]]).split(".")[0].strip().zfill(5)+school
        try: enroll=int(float(row[cols["E"]]))
        except: enroll=""
        try: upp=round(float(row[cols["U"]])*100,2)
        except: upp=""
        flag=row[cols["F"]]; flag=(str(flag).strip() if flag is not None else "")
        out.append([cds,yr,enroll,upp,flag]); n+=1
    sys.stderr.write(f"{os.path.basename(fp)} ({yr}): cols={cols} -> {n}\n")
with open("build_outputs/upp_lcff.csv","w",newline="") as f:
    w=csv.writer(f); w.writerow(["cds14","cupc_year","enroll_9_12","upp_pct","lcff_plus_flag"]); w.writerows(out)
print("upp rows=",len(out))
